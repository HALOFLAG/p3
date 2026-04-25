# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pdfplumber>=0.11",
#   "openpyxl>=3.1",
# ]
# ///
"""
Extract all PDFs + trpg.xlsx in nobinobiTRPG參考資料/ into a single Markdown
reference file. Intermediate .txt files live in a TemporaryDirectory that is
auto-removed on exit.

Run: uv run tools/extract_reference.py
"""

from __future__ import annotations

import re
import sys
import tempfile
import traceback
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from card_lib import extract_cards_from_pdf


ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT / "nobinobiTRPG參考資料"
XLSX_NAME = "trpg.xlsx"
OUT_MD = ROOT / "nobinobiTRPG_參考資料.md"

MAX_CELL = 200  # truncate overly long excel cells
MAX_ROWS_PER_SHEET = 20000  # safety cap


@dataclass
class PdfStat:
    name: str
    pages: int = 0
    chars: int = 0
    text_layer: bool = True
    cards: int = 0
    dropped: int = 0
    error: str | None = None


@dataclass
class SheetStat:
    name: str
    rows: int = 0
    cols: int = 0
    truncated: bool = False
    error: str | None = None


@dataclass
class Report:
    pdfs: list[PdfStat] = field(default_factory=list)
    sheets: list[SheetStat] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------- sorting --------------------------------------------------------

def sort_key(p: Path) -> tuple:
    """Sort PDFs as: 起始卡 → 事件卡 → 事件卡 X → 結局卡. 母板 last within group."""
    name = p.stem
    if name.startswith("起始卡"):
        group = 0
    elif name.startswith("事件卡 X"):
        group = 2
    elif name.startswith("事件卡"):
        group = 1
    elif name.startswith("結局卡"):
        group = 3
    else:
        group = 9

    is_template = "母板" in name
    m = re.search(r"(\d+)", name)
    num = int(m.group(1)) if m else 999
    ex = 1 if "EX" in name or "ex" in name else 0
    return (group, 1 if is_template else 0, num, ex, name)


# ---------- pdf ------------------------------------------------------------

def _format_card_md(card: list[str]) -> str:
    title = card[0].strip()
    body = "\n".join(ln.strip() for ln in card[1:] if ln.strip())
    return f"### {title}\n\n{body}\n"


def extract_pdf(pdf_path: Path, txt_out: Path) -> PdfStat:
    stat = PdfStat(name=pdf_path.name)
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            stat.pages = len(pdf.pages)
        kept, dropped = extract_cards_from_pdf(pdf_path)
        stat.cards = len(kept)
        stat.dropped = dropped
        if not kept and dropped == 0:
            stat.text_layer = False
            body = "[無文字層 — PDF 可能為圖片掃描, 需 OCR]"
        elif not kept:
            body = "[所有卡片片段皆被判定為無法分辨, 已略過]"
        else:
            body = "\n".join(_format_card_md(c) for c in kept)
        stat.chars = len(body)
        txt_out.write_text(body, encoding="utf-8")
    except Exception as exc:
        stat.error = f"{type(exc).__name__}: {exc}"
    return stat


# ---------- xlsx -----------------------------------------------------------

def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # openpyxl loses precision on some ints stored as float; reclaim
        if v.is_integer():
            v = int(v)
        elif abs(v - round(v)) < 1e-6:
            v = int(round(v))
        else:
            v = round(v, 6)
    s = str(v).replace("\r", " ").replace("\n", " ").replace("|", "\\|").strip()
    if len(s) > MAX_CELL:
        s = s[:MAX_CELL] + "…"
    return s


def sheet_to_md(ws) -> tuple[str, SheetStat]:
    stat = SheetStat(name=ws.title)
    # collect rows. Stop after 30 consecutive rows where the first column is
    # empty — handles xlsx sheets with auto-increment formulas filling column B
    # down to the sheet's 1_048_576 row limit.
    rows: list[list[str]] = []
    max_cols = 0
    empty_title_run = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= MAX_ROWS_PER_SHEET:
            stat.truncated = True
            break
        cells = [_cell(c) for c in row]
        if not any(cells):
            continue
        if not cells[0]:
            empty_title_run += 1
            if empty_title_run >= 30:
                break
            continue
        empty_title_run = 0
        rows.append(cells)
        max_cols = max(max_cols, len(cells))

    if not rows:
        return "_(sheet 為空)_\n", stat

    # pad rows to equal width
    for r in rows:
        r.extend([""] * (max_cols - len(r)))

    stat.rows = len(rows)
    stat.cols = max_cols

    # use a synthetic header so no data row is consumed — these sheets have no
    # real header row, the first data entry starts at row 1
    header = [f"欄{chr(ord('A') + i)}" for i in range(max_cols)]
    md = ["| " + " | ".join(header) + " |",
          "| " + " | ".join(["---"] * max_cols) + " |"]
    for r in rows:
        md.append("| " + " | ".join(r) + " |")
    if stat.truncated:
        md.append(f"\n_(超過 {MAX_ROWS_PER_SHEET} 列, 已截斷)_")
    return "\n".join(md) + "\n", stat


# ---------- assemble -------------------------------------------------------

def main() -> int:
    if not SRC_DIR.is_dir():
        print(f"ERROR: source dir not found: {SRC_DIR}", file=sys.stderr)
        return 2

    report = Report()
    out_parts: list[str] = []
    out_parts.append("# nobinobiTRPG 參考資料\n")
    out_parts.append(
        "_由 `tools/extract_reference.py` 自 "
        f"`{SRC_DIR.name}/` 批次提取。_\n"
    )

    pdf_files = sorted(SRC_DIR.glob("*.pdf"), key=sort_key)

    with tempfile.TemporaryDirectory(prefix="trpg_ref_") as tmp:
        tmp_dir = Path(tmp)
        out_parts.append("\n# 一、卡片 (PDF)\n")

        for pdf in pdf_files:
            print(f"[pdf] {pdf.name}")
            txt = tmp_dir / f"{pdf.stem}.txt"
            stat = extract_pdf(pdf, txt)
            report.pdfs.append(stat)
            out_parts.append(f"\n## {pdf.stem}\n")
            if stat.error:
                out_parts.append(f"_[提取失敗: {stat.error}]_\n")
                report.errors.append(f"{pdf.name}: {stat.error}")
                continue
            body = txt.read_text(encoding="utf-8") if txt.exists() else ""
            out_parts.append(body)
            if stat.dropped:
                out_parts.append(
                    f"\n_(已略過 {stat.dropped} 個無法清楚分辨的片段)_\n"
                )

        # xlsx
        xlsx = SRC_DIR / XLSX_NAME
        out_parts.append("\n# 二、Excel (trpg.xlsx)\n")
        if xlsx.is_file():
            try:
                print(f"[xlsx] {xlsx.name}")
                wb = load_workbook(xlsx, data_only=True, read_only=True)
                for sname in wb.sheetnames:
                    print(f"  sheet: {sname}")
                    ws = wb[sname]
                    md, sstat = sheet_to_md(ws)
                    report.sheets.append(sstat)
                    out_parts.append(f"\n## Excel — {sname}\n\n{md}")
                wb.close()
            except Exception as exc:
                msg = f"{type(exc).__name__}: {exc}"
                out_parts.append(f"\n_[xlsx 載入失敗: {msg}]_\n")
                report.errors.append(f"{xlsx.name}: {msg}")
                traceback.print_exc()
        else:
            out_parts.append(f"_(未找到 {XLSX_NAME})_\n")

        # stats
        out_parts.append("\n# 三、提取統計\n")
        out_parts.append("\n## PDF\n")
        out_parts.append("| 檔名 | 頁 | 卡片 | 略過 | 字元 | 文字層 | 錯誤 |")
        out_parts.append("| --- | ---: | ---: | ---: | ---: | :---: | --- |")
        for s in report.pdfs:
            layer = "○" if s.text_layer else "✗"
            err = s.error or ""
            out_parts.append(
                f"| {s.name} | {s.pages} | {s.cards} | {s.dropped} "
                f"| {s.chars} | {layer} | {err} |"
            )

        out_parts.append("\n## Excel sheets\n")
        out_parts.append("| Sheet | 列 | 欄 | 截斷 | 錯誤 |")
        out_parts.append("| --- | ---: | ---: | :---: | --- |")
        for s in report.sheets:
            trunc = "✗" if s.truncated else ""
            err = s.error or ""
            out_parts.append(
                f"| {s.name} | {s.rows} | {s.cols} | {trunc} | {err} |"
            )

        if report.errors:
            out_parts.append("\n## 失敗清單\n")
            for e in report.errors:
                out_parts.append(f"- {e}")

    OUT_MD.write_text("\n".join(out_parts) + "\n", encoding="utf-8")

    # console summary
    print()
    print(f"PDFs processed : {len(report.pdfs)} "
          f"(no text layer: {sum(1 for s in report.pdfs if not s.text_layer)}, "
          f"errors: {sum(1 for s in report.pdfs if s.error)})")
    print(f"Excel sheets   : {len(report.sheets)}")
    if report.errors:
        print(f"Errors         : {len(report.errors)}")
    print(f"Output         : {OUT_MD}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

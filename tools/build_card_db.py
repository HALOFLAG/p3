# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pdfplumber>=0.11",
#   "openpyxl>=3.1",
# ]
# ///
"""
Build a unified card database (xlsx + JSON) from:
  - nobinobiTRPG參考資料/*.pdf  (30 PDFs → 154 cards after cleaning)
  - nobinobiTRPG參考資料/trpg.xlsx  (9 sheets → ~132 rows)

Schema (6 columns):
  類別 | 來源 | 編號 | 標題 | 故事 | 判定

Deduplication: 母板 cards whose (title + story) strongly match a non-母板 card
are dropped (templates duplicate the live deck).

Outputs: cards.xlsx (multi-sheet by category) + cards.json (flat records).

Run: uv run tools/build_card_db.py
"""

from __future__ import annotations

import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from card_lib import extract_cards_from_pdf, split_card_body


ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT / "nobinobiTRPG參考資料"
XLSX_SRC = SRC_DIR / "trpg.xlsx"
OUT_XLSX = ROOT / "cards.xlsx"
OUT_JSON = ROOT / "cards.json"

# 母板 dedup threshold (story similarity ratio, title must match exactly)
DEDUP_TITLE_EXACT = True
DEDUP_STORY_THRESHOLD = 0.85

# fixed column order + widths
COLUMNS = ["類別", "來源", "編號", "標題", "故事", "判定"]
COL_WIDTHS = [10, 22, 12, 24, 70, 55]

# PDF category mapping
PDF_CATEGORIES = [
    ("起始卡母板", "母板"),       # must be checked before 起始卡
    ("事件卡 母板", "母板"),
    ("起始卡", "起始卡"),
    ("事件卡 X", "事件卡X"),
    ("事件卡", "事件卡"),
    ("結局卡", "結局卡"),
]


@dataclass
class Card:
    category: str
    source: str
    number: str
    title: str
    story: str
    check: str

    def as_row(self) -> list[str]:
        return [self.category, self.source, self.number,
                self.title, self.story, self.check]

    def to_dict(self) -> dict:
        return {
            "category": self.category,
            "source": self.source,
            "number": self.number,
            "title": self.title,
            "story": self.story,
            "check": self.check,
        }


# ---------- ingest ---------------------------------------------------------

def _categorize_pdf(stem: str) -> str:
    for prefix, cat in PDF_CATEGORIES:
        if stem.startswith(prefix):
            return cat
    return "其他"


def _pdf_sort_key(p: Path) -> tuple:
    name = p.stem
    if name.startswith("起始卡"):
        g = 0
    elif name.startswith("事件卡 X"):
        g = 2
    elif name.startswith("事件卡"):
        g = 1
    elif name.startswith("結局卡"):
        g = 3
    else:
        g = 9
    tmpl = 1 if "母板" in name else 0
    m = re.search(r"(\d+)", name)
    num = int(m.group(1)) if m else 999
    ex = 1 if ("EX" in name or "ex" in name) else 0
    return (g, tmpl, num, ex, name)


def ingest_pdfs() -> tuple[list[Card], int]:
    cards: list[Card] = []
    total_dropped = 0
    per_cat_counter: dict[str, int] = {}
    for pdf in sorted(SRC_DIR.glob("*.pdf"), key=_pdf_sort_key):
        cat = _categorize_pdf(pdf.stem)
        kept, dropped = extract_cards_from_pdf(pdf)
        total_dropped += dropped
        for raw in kept:
            title = raw[0].strip()
            story, check = split_card_body(raw)
            per_cat_counter[cat] = per_cat_counter.get(cat, 0) + 1
            num = f"{cat[:2]}-{per_cat_counter[cat]:03d}"
            cards.append(Card(
                category=cat,
                source=pdf.name,
                number=num,
                title=title,
                story=story,
                check=check,
            ))
    return cards, total_dropped


def _xls_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            v = int(v)
        elif abs(v - round(v)) < 1e-6:
            v = int(round(v))
        else:
            v = round(v, 6)
    return str(v).strip()


def ingest_xlsx() -> list[Card]:
    """Each sheet in trpg.xlsx becomes its own category. Column layout:
      A = title, B = number, C = story, D = check (optional; 3-col sheets
      collapse story+check into column C)."""
    cards: list[Card] = []
    if not XLSX_SRC.is_file():
        return cards
    wb = load_workbook(XLSX_SRC, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        per_sheet_idx = 0
        empty_run = 0
        for row in ws.iter_rows(values_only=True):
            cells = [_xls_cell(c) for c in row]
            if not any(cells):
                continue
            # drop rows without a title (handles the phantom-row issue in HB)
            title = cells[0] if cells else ""
            if not title:
                empty_run += 1
                if empty_run >= 30:
                    break
                continue
            empty_run = 0
            number = cells[1] if len(cells) > 1 else ""
            story = cells[2] if len(cells) > 2 else ""
            check = cells[3] if len(cells) > 3 else ""
            per_sheet_idx += 1
            cat = sheet_name.strip() or "xlsx"
            cards.append(Card(
                category=cat,
                source=f"trpg.xlsx:{sheet_name.strip()}",
                number=number or f"{cat}-{per_sheet_idx:03d}",
                title=title,
                story=story,
                check=check,
            ))
    wb.close()
    return cards


# ---------- dedup 母板 -----------------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"\s+", "", s).strip()


def dedup_templates(cards: list[Card]) -> tuple[list[Card], list[Card]]:
    """Drop 母板 cards that duplicate a non-母板 card.

    Returns (kept, dropped)."""
    non_templates = [c for c in cards if c.category != "母板"]
    kept: list[Card] = []
    dropped: list[Card] = []
    for c in cards:
        if c.category != "母板":
            kept.append(c)
            continue
        is_dup = False
        c_title = _norm(c.title)
        c_story = _norm(c.story)
        for other in non_templates:
            if DEDUP_TITLE_EXACT and _norm(other.title) != c_title:
                continue
            sim = SequenceMatcher(None, c_story, _norm(other.story)).ratio()
            if sim >= DEDUP_STORY_THRESHOLD:
                is_dup = True
                break
        if is_dup:
            dropped.append(c)
        else:
            kept.append(c)
    return kept, dropped


# ---------- output ---------------------------------------------------------

def _write_header_row(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="E8ECEF")
    font = Font(bold=True)
    for i, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_card_rows(ws, cards: list[Card], start_row: int = 2) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    for r, c in enumerate(cards, start_row):
        for col_i, val in enumerate(c.as_row(), 1):
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.alignment = wrap
    ws.freeze_panes = "A2"


def write_xlsx(cards: list[Card], stats: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # overview sheet first
    ov = wb.create_sheet("概覽")
    ov.append(["項目", "數量"])
    ov["A1"].font = Font(bold=True)
    ov["B1"].font = Font(bold=True)
    ov.column_dimensions["A"].width = 28
    ov.column_dimensions["B"].width = 12
    rows = [
        ("卡片總數（去重後）", len(cards)),
        ("PDF 片段略過數", stats.get("dropped_pdf", 0)),
        ("母板去重剔除數", stats.get("dropped_template", 0)),
        ("", ""),
        ("— 依類別 —", ""),
    ]
    counter: dict[str, int] = {}
    for c in cards:
        counter[c.category] = counter.get(c.category, 0) + 1
    # stable order: PDF categories first, then xlsx sheet codes
    pdf_cats = ["起始卡", "事件卡", "事件卡X", "結局卡", "母板"]
    xlsx_cats = sorted(k for k in counter if k not in pdf_cats)
    for cat in pdf_cats + xlsx_cats:
        if cat in counter:
            rows.append((cat, counter[cat]))
    for r in rows:
        ov.append(r)
    ov.append(("", ""))
    ov.append(("產生時間", str(date.today())))

    # one sheet per category
    for cat in pdf_cats + xlsx_cats:
        subset = [c for c in cards if c.category == cat]
        if not subset:
            continue
        # excel sheet name cannot exceed 31 chars or contain :\/?*[]
        safe = re.sub(r"[:\\/?*\[\]]", "_", cat)[:31]
        ws = wb.create_sheet(safe)
        _write_header_row(ws)
        _write_card_rows(ws, subset)

    wb.save(OUT_XLSX)


def write_json(cards: list[Card], stats: dict) -> None:
    payload = {
        "generated": str(date.today()),
        "schema": COLUMNS,
        "total": len(cards),
        "stats": stats,
        "cards": [c.to_dict() for c in cards],
    }
    OUT_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ---------- main -----------------------------------------------------------

def main() -> int:
    if not SRC_DIR.is_dir():
        print(f"ERROR: {SRC_DIR} not found", file=sys.stderr)
        return 2

    print("[1/4] ingesting PDFs…")
    pdf_cards, pdf_dropped = ingest_pdfs()
    print(f"      {len(pdf_cards)} cards, {pdf_dropped} fragments skipped")

    print("[2/4] ingesting trpg.xlsx…")
    xlsx_cards = ingest_xlsx()
    print(f"      {len(xlsx_cards)} cards across "
          f"{len({c.category for c in xlsx_cards})} sheets")

    all_cards = pdf_cards + xlsx_cards
    print(f"      total before dedup: {len(all_cards)}")

    print("[3/4] deduping 母板…")
    kept, dropped = dedup_templates(all_cards)
    print(f"      kept {len(kept)}, dropped {len(dropped)} duplicate templates")
    for d in dropped:
        print(f"         · {d.source} → {d.title}")

    stats = {
        "total_before_dedup": len(all_cards),
        "dropped_pdf": pdf_dropped,
        "dropped_template": len(dropped),
        "dropped_template_list": [
            {"source": d.source, "title": d.title} for d in dropped
        ],
    }

    print("[4/4] writing outputs…")
    write_xlsx(kept, stats)
    write_json(kept, stats)

    print()
    print(f"xlsx  : {OUT_XLSX}")
    print(f"json  : {OUT_JSON}")
    print(f"total : {len(kept)} cards")
    return 0


if __name__ == "__main__":
    sys.exit(main())
